from __future__ import annotations

import os
import re
import tempfile
import uuid
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from sqlalchemy import text

from sql_agent.query_utils import (
    _mask_sql_literals_identifiers_and_comments,
    validate_readonly_select_sql,
)


EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_CELL_CHARACTERS = 32_767
SQL_LINE_BREAK_PATTERN = re.compile(
    r",|\b(?:UNION\s+ALL|ORDER\s+BY|GROUP\s+BY|LEFT\s+OUTER\s+JOIN|"
    r"RIGHT\s+OUTER\s+JOIN|FULL\s+OUTER\s+JOIN|INNER\s+JOIN|LEFT\s+JOIN|"
    r"RIGHT\s+JOIN|FULL\s+JOIN|CROSS\s+JOIN|SELECT|FROM|WHERE|HAVING|JOIN|ON|WITH)\b",
    flags=re.IGNORECASE,
)
SQL_CLAUSE_PREFIX_PATTERN = re.compile(
    r"^(?:WITH|SELECT|FROM|WHERE|HAVING|UNION\s+ALL|ORDER\s+BY|GROUP\s+BY|"
    r"LEFT\s+OUTER\s+JOIN|RIGHT\s+OUTER\s+JOIN|FULL\s+OUTER\s+JOIN|"
    r"INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|FULL\s+JOIN|CROSS\s+JOIN|JOIN)\b",
    flags=re.IGNORECASE,
)
SQL_INDENT = "    "


def export_sql_to_excel(engine, sql: str) -> tuple[Path, int]:
    """Execute a read-only query and write its rows to a temporary XLSX file."""
    validate_readonly_select_sql(sql)
    file_handle, raw_path = tempfile.mkstemp(prefix="atlas_sql_export_", suffix=".xlsx")
    # mkstemp's descriptor must be closed before openpyxl opens the path on Windows.
    os.close(file_handle)
    path = Path(raw_path)
    path.unlink(missing_ok=True)
    workbook = Workbook(write_only=True)
    row_count = 0

    try:
        with engine.connect() as connection:
            result = connection.execution_options(stream_results=True).execute(text(sql))
            columns = [str(column) for column in result.keys()]
            sheet_index = 1
            sheet = workbook.create_sheet(_sheet_title(sheet_index))
            sheet.append(columns)
            rows_in_sheet = 1

            for raw_row in result:
                if rows_in_sheet >= EXCEL_MAX_ROWS:
                    sheet_index += 1
                    sheet = workbook.create_sheet(_sheet_title(sheet_index))
                    sheet.append(columns)
                    rows_in_sheet = 1
                sheet.append([_excel_value(value) for value in raw_row])
                rows_in_sheet += 1
                row_count += 1

        _append_sql_sheet(workbook, sql)
        workbook.save(path)
        return path, row_count
    except Exception:
        path.unlink(missing_ok=True)
        raise
    finally:
        workbook.close()


def _sheet_title(index: int) -> str:
    return "Данные" if index == 1 else f"Данные {index}"


def _append_sql_sheet(workbook: Workbook, sql: str) -> None:
    sheet = workbook.create_sheet("SQL")
    sheet.append(["SQL скрипт"])
    formatted_sql = _format_sql_for_excel(ILLEGAL_CHARACTERS_RE.sub("", sql))
    for line in formatted_sql.splitlines():
        for start in range(0, len(line), EXCEL_MAX_CELL_CHARACTERS):
            sheet.append([line[start : start + EXCEL_MAX_CELL_CHARACTERS]])


def _format_sql_for_excel(sql: str) -> str:
    masked_sql = _mask_sql_literals_identifiers_and_comments(sql)
    parts: list[str] = []
    cursor = 0
    for match in SQL_LINE_BREAK_PATTERN.finditer(masked_sql):
        preceding_text = sql[cursor : match.start()].rstrip()
        if preceding_text:
            parts.append(preceding_text)
        if match.group() == ",":
            parts.extend((",", "\n"))
            cursor = match.end()
            continue
        if parts and not parts[-1].endswith(("\n", "\r")):
            parts.append("\n")
        parts.append(sql[match.start() : match.end()])
        cursor = match.end()
    parts.append(sql[cursor:])
    formatted_sql = "\n".join(
        line.strip() for line in "".join(parts).splitlines() if line.strip()
    )
    return _indent_formatted_sql(formatted_sql)


def _indent_formatted_sql(sql: str) -> str:
    indented_lines: list[str] = []
    parenthesis_depth = 0
    previous_line_ends_with_comma = False

    for line in sql.splitlines():
        stripped_line = line.strip()
        masked_line = _mask_sql_literals_identifiers_and_comments(stripped_line)
        leading_closings = len(masked_line) - len(masked_line.lstrip(")"))
        line_depth = max(parenthesis_depth - leading_closings, 0)
        extra_indent = 0
        if re.match(r"^ON\b", stripped_line, flags=re.IGNORECASE):
            extra_indent = 1
        elif previous_line_ends_with_comma and not SQL_CLAUSE_PREFIX_PATTERN.match(
            stripped_line
        ):
            extra_indent = 1

        indented_lines.append(
            f"{SQL_INDENT * (line_depth + extra_indent)}{stripped_line}"
        )
        parenthesis_depth = max(
            parenthesis_depth + masked_line.count("(") - masked_line.count(")"),
            0,
        )
        previous_line_ends_with_comma = stripped_line.endswith(",")

    return "\n".join(indented_lines)


def _excel_value(value):
    if value is None:
        return None
    if isinstance(value, (bytes, bytearray, memoryview)):
        raw_value = bytes(value)
        value = str(uuid.UUID(bytes_le=raw_value)) if len(raw_value) == 16 else "0x" + raw_value.hex()
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if value.startswith(("=", "+", "-", "@")):
            value = "'" + value
        return value[:EXCEL_MAX_CELL_CHARACTERS]
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.isoformat()
    if isinstance(value, (bool, int, float, Decimal, date, datetime, time, timedelta)):
        return value
    return str(value)[:EXCEL_MAX_CELL_CHARACTERS]

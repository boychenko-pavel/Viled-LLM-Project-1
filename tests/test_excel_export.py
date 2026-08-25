from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from sql_agent.excel_export import _format_sql_for_excel, export_sql_to_excel
from sql_agent.intents import QueryIntent
from sql_agent.memory import SqlAgentMemoryRepository
from sql_agent.service import SqlAgentService
from sql_agent.web import _excel_export_filename


def test_excel_export_streams_query_rows_to_workbook() -> None:
    engine = create_engine("sqlite://")
    with engine.begin() as connection:
        connection.execute(text("CREATE TABLE sample (id INTEGER, note TEXT)"))
        connection.execute(
            text("INSERT INTO sample (id, note) VALUES (1, 'alpha'), (2, '=1+1')")
        )

    path, row_count = export_sql_to_excel(
        engine,
        "SELECT id, note FROM sample ORDER BY id",
    )
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        rows = list(workbook["Данные"].iter_rows(values_only=True))
        sql_rows = list(workbook["SQL"].iter_rows(values_only=True))
        workbook.close()
    finally:
        path.unlink(missing_ok=True)

    assert row_count == 2
    assert rows == [("id", "note"), (1, "alpha"), (2, "'=1+1")]
    assert sql_rows == [
        ("SQL скрипт",),
        ("SELECT id,",),
        ("    note",),
        ("FROM sample",),
        ("ORDER BY id",),
    ]


def test_excel_sql_formatting_does_not_split_quoted_values() -> None:
    assert _format_sql_for_excel(
        "SELECT 'FROM, WHERE' AS note, id FROM sample WHERE note = 'ORDER BY'"
    ).splitlines() == [
        "SELECT 'FROM, WHERE' AS note,",
        "    id",
        "FROM sample",
        "WHERE note = 'ORDER BY'",
    ]


def test_excel_sql_formatting_indents_nested_queries_and_join_conditions() -> None:
    formatted_sql = _format_sql_for_excel(
        "WITH scope AS (SELECT id, name FROM sample) "
        "SELECT scope.id FROM scope INNER JOIN other ON scope.id = other.id"
    )

    assert formatted_sql.splitlines() == [
        "WITH scope AS (",
        "    SELECT id,",
        "        name",
        "    FROM sample)",
        "SELECT scope.id",
        "FROM scope",
        "INNER JOIN other",
        "    ON scope.id = other.id",
    ]


def test_export_sql_generation_removes_web_row_limit(tmp_path) -> None:
    class Parser:
        def parse(self, question, memory):
            return QueryIntent(operation="select", domain="sales", limit=100)

    class Builder:
        def execute(self, db, intent, on_sql_ready):
            assert intent.limit is None
            on_sql_ready("SELECT id FROM sample")

    class Connector:
        def build_engine(self):
            return object()

    service = SqlAgentService(
        memory_repository=SqlAgentMemoryRepository(tmp_path / "memory.json"),
        database_connector=Connector(),
        intent_parser=Parser(),
        sql_builder=Builder(),
    )

    assert service.build_export_sql("все строки продаж") == "SELECT id FROM sample"


def test_export_sql_for_without_top_sales_request_has_no_top(tmp_path) -> None:
    class Connector:
        def build_engine(self):
            return object()

    service = SqlAgentService(
        memory_repository=SqlAgentMemoryRepository(tmp_path / "memory.json"),
        database_connector=Connector(),
    )

    sql = service.build_export_sql(
        "продажи бренд Cartier август 2026 без лимита"
    )

    assert "TOP " not in sql.upper()
    assert "dim.[brand] = 'Cartier'" in sql
    assert "fact.[sale_date] BETWEEN '2026-08-01' AND '2026-08-31'" in sql


def test_unbounded_web_warning_includes_export_sql(tmp_path) -> None:
    class Connector:
        def build_engine(self):
            return object()

    service = SqlAgentService(
        memory_repository=SqlAgentMemoryRepository(tmp_path / "memory.json"),
        database_connector=Connector(),
    )

    response = service.ask_database(
        "продажи бренд Cartier август 2026 без лимита"
    )

    assert response.startswith("Безлимитный вывод строк в веб-чате отключён")
    assert "\n\nSQL:\n" in response
    assert "dim.[brand] = 'Cartier'" in response
    assert "fact.[sale_date] BETWEEN '2026-08-01' AND '2026-08-31'" in response
    assert "TOP " not in response.upper()


def test_excel_export_filename_contains_creation_timestamp() -> None:
    assert _excel_export_filename(datetime(2026, 8, 24, 14, 5, 9)) == (
        "viled_atlas_sql_agent 2026-08-24 14-05-09.xlsx"
    )
